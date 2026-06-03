import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_xlsx(input_json_path, output_xlsx_path):
    with open(input_json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    headers = data['headers']
    rows = data['rows']
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AWB Export"
    
    # Enable grid lines visibility explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Professional styling definitions
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Navy Blue
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="000000")
    
    thin_border_side = Side(style='thin', color='D9D9D9')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_all
    
    # Row height for header
    ws.row_dimensions[1].height = 28
    
    # Write data
    for row_num, row_data in enumerate(rows, 2):
        ws.row_dimensions[row_num].height = 20
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.font = data_font
            cell.border = border_all
            
            # Format and align dynamically based on header name
            header_name = headers[col_num - 1]
            
            # Numeric alignment and formatting
            if any(term in header_name for term in ["WEIGHT", "LENGTH", "WIDTH", "HEIGHT", "PIECES", "RATE", "FREIGHT", "SURCHARGE", "AMOUNT", "CHARGES", "TAXABLE", "DDC", "SSC", "BTC", "FLC", "DOC", "TSP", "TCP", "ADO", "FEES", "STORAGE", "HANDLING"]):
                cell.alignment = align_right
                # Format currency
                if any(term in header_name for term in ["RATE", "FREIGHT", "AMOUNT", "CHARGES", "TAXABLE", "DDC", "SSC", "BTC", "FLC", "DOC", "TSP", "TCP", "ADO", "FEES", "STORAGE", "HANDLING"]):
                    cell.number_format = '[$₹-380A] #,##0.00' # Indian Rupee Format
                elif "PIECES" in header_name:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.0'
            elif any(term in header_name for term in ["DATE", "TIME", "SR NO", "AWB NO", "DOCKET NO", "GST APPLIED", "GSTIN", "PAN", "SAC CODE"]):
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    # Auto-fit columns with safety margin
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    # Enable Auto Filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    
    wb.save(output_xlsx_path)

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python generate_xlsx.py <input_json> <output_xlsx>")
        sys.exit(1)
    generate_xlsx(sys.argv[1], sys.argv[2])
